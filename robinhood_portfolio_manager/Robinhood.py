################################################################################################################
# A python CLI that allows a Robinhood user to easily allocate and divide his capital on different stocks.
################################################################################################################
# Author: Mina Messiah
# Copyright: Copyright 2021, robinhood-portfolio-manager
# License: MIT License
# Version: 0.0b9
# Date: 02-25-2021
# Email: mena.sb.109@gmail.com
# URLs: pypi.org/project/robinhood-portfolio-manager &
# github.com/MinaMessiah/robinhood_portfolio_manager
# Status: In development
################################################################################################################


import os
import csv
import time
import pyotp
import logging
import pathlib
import platform
import openpyxl
import openpyxl.styles
import robin_stocks.robinhood as rh


class Robinhood(object):

    def __init__(self, **kwargs):
        self.logger = self._configure_logger()

        username = kwargs["username"] if "username" in kwargs else os.getenv("ROBINHOOD_USERNAME")
        password = kwargs["password"] if "password" in kwargs else os.getenv("ROBINHOOD_PASSWORD")
        totp = kwargs["totp"] if "totp" in kwargs else os.getenv("ROBINHOOD_TOTP")

        if not all([username, password, totp]):
            message = "One or more environment variable(s) are not set. set ROBINHOOD_USERNAME, ROBINHOOD_PASSWORD, ROBINHOOD_TOTP variables or provide credentials through command line arguments."
            self._log(logging.ERROR, message)
            raise Exception(message)

        totp = pyotp.TOTP(totp).now()
        self.login = rh.authentication.login(username, password, store_session=False, mfa_code=totp)


    def _configure_logger(self):
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)

        file_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        stream_formatter = logging.Formatter("%(message)s")
        
        if platform.system().lower() == "windows":
            file_handler = logging.FileHandler(pathlib.Path.home().joinpath("Documents", "Robinhood.log"))
        else:
            file_handler = logging.FileHandler(pathlib.Path.home().joinpath("Robinhood.log"))
        
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(file_formatter)

        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(logging.DEBUG)
        stream_handler.setFormatter(stream_formatter)
        
        logger.addHandler(file_handler)
        logger.addHandler(stream_handler)
        return logger


    def _log(self, level, message, sim=False):
        if sim:
            self.logger.log(level, "SIMULATION: " + message)
        else:
            self.logger.log(level, message)


    def _check_list(self):
        if platform.system().lower() == "windows":
            xlsx_file = pathlib.Path.home().joinpath("Documents", "Robinhood.xlsx")
        else:
            xlsx_file = pathlib.Path.home().joinpath("Robinhood.xlsx")

        if platform.system().lower() == "windows":
            csv_file = pathlib.Path.home().joinpath("Documents", "Robinhood.csv")
        else:
            csv_file = pathlib.Path.home().joinpath("Robinhood.csv")

        if xlsx_file.is_file():
            self._check_excel_list(xlsx_file)
        elif csv_file.is_file():
            self._check_csv_list(csv_file)
        else:
            message = "Excel/CSV file not found. Run generate_excel_file()."
            self._log(logging.ERROR, message)
            raise Exception(message)


    def _check_excel_list(self, xlsx_file):
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        max_row = len(sheet["A"]) - 1
        
        portfolio = rh.account.get_watchlist_by_name("Portfolio")
        if not portfolio:
            message = "Watchlist not found on Robinhood. Create 'Portfolio' watchlist."
            self._log(logging.ERROR, message)
            raise Exception(message)

        excel_list = [row[0].value for row in sheet.iter_rows(min_row=2, max_row=max_row)]
        portfolio_list = [instrument["symbol"] for instrument in portfolio["results"]]
        diff = (set(excel_list) - set(portfolio_list)) | (set(portfolio_list) - set(excel_list))

        total_percent = sum([float(row[1].value) for row in sheet.iter_rows(min_row=2, max_row=max_row)])

        if round(total_percent, 2) != 1:
            message = "Total percentage doesn't add up to 100%. Check excel file."
            self._log(logging.ERROR, message)
            raise Exception(message)
        elif diff:
            message = "Excel file and Robinhood watchlist not in sync."
            self._log(logging.ERROR, message)
            raise Exception(message)
        return True


    def _check_csv_list(self, csv_file):
        _list = []

        with open(csv_file, 'r') as file:
            data = csv.reader(file)
            for row in data:
                if row:
                    _list.append(row)
        
        portfolio = rh.account.get_watchlist_by_name("Portfolio")
        
        if not portfolio:
            message = "Watchlist not found on Robinhood. Create 'Portfolio' watchlist."
            self._log(logging.ERROR, message)
            raise Exception(message)

        csv_list = [row[0] for row in _list[1:]]
        portfolio_list = [instrument["symbol"] for instrument in portfolio["results"]]
        diff = (set(csv_list) - set(portfolio_list)) | (set(portfolio_list) - set(csv_list))

        total_percent = sum([float(row[1]) for row in _list[1:]])

        if round(total_percent, 2) != 100:
            print(total_percent)
            message = "Total percentage doesn't add up to 100%. Check CSV file."
            self._log(logging.ERROR, message)
            raise Exception(message)
        elif diff:
            message = "CSV file and Robinhood watchlist not in sync."
            self._log(logging.ERROR, message)
            raise Exception(message)
        return True


    def _get_portfolio_value(self):
        buying_power = float(rh.profiles.load_account_profile(info="buying_power"))
        equity = rh.profiles.load_portfolio_profile(info="extended_hours_portfolio_equity")

        if not equity:
            equity = rh.profiles.load_portfolio_profile(info="equity")

        return float(equity) + buying_power


    def _check_if_collateral(self):
        for instrument in self.get_current_investments.values():
            if instrument["collateral_shares"]:
                return True

        return False


    def generate_excel_file(self):
        if platform.system().lower() == "windows":
            xlsx_file = pathlib.Path.home().joinpath("Documents", "Robinhood.xlsx")
        else:
            xlsx_file = pathlib.Path.home().joinpath("Robinhood.xlsx")
            
        if xlsx_file.is_file():
            message = "Excel file found. Delete file to regenerate."
            self._log(logging.ERROR, message)
            raise Exception(message)

        portfolio_list = rh.account.get_watchlist_by_name("Portfolio")["results"]
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Portfolio"

        worksheet["A1"] = "Symbol"
        worksheet["B1"] = "Percentage"
        worksheet["A1"].font = openpyxl.styles.Font(bold=True)
        worksheet["B1"].font = openpyxl.styles.Font(bold=True)
        worksheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
        worksheet["B1"].alignment = openpyxl.styles.Alignment(horizontal="center")
        worksheet.column_dimensions["A"].width = 12
        worksheet.column_dimensions["B"].width = 12

        row = 2
        for instrument in portfolio_list:
            worksheet.cell(row=row, column=1).value = instrument["symbol"]
            worksheet.cell(row=row, column=2).value = 0.0
            worksheet.cell(row=row, column=2).number_format = '0.00%'
            row += 1

        worksheet.cell(row=row, column=1).value = "Total"
        worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
        worksheet.cell(row=row, column=2).value = f"=SUM(B1:B{row-1})"
        worksheet.cell(row=row, column=2).number_format = '0.00%'
        worksheet.cell(row=row, column=2).font = openpyxl.styles.Font(bold=True, color="006400")
        workbook.save(xlsx_file)

        self._log(logging.INFO, "Excel file generated. Configure percentages for each stock.")


    def generate_csv_file(self):
        if platform.system().lower() == "windows":
            csv_file = pathlib.Path.home().joinpath("Documents", "Robinhood.csv")
        else:
            csv_file = pathlib.Path.home().joinpath("Robinhood.csv")
            
        if csv_file.is_file():
            message = "CSV file found. Delete file to regenerate."
            self._log(logging.ERROR, message)
            raise Exception(message)

        portfolio_list = rh.account.get_watchlist_by_name("Portfolio")["results"]

        rows = [[instrument["symbol"], "0.00"] for instrument in portfolio_list]
        with open(csv_file, "w") as file:
            writer = csv.writer(file)
            writer.writerow(["Symbol", "Percentage"])
            writer.writerows(rows)
        
        self._log(logging.INFO, "CSV file generated. Configure percentages for each stock.")


    def get_current_investments(self):
        market_value = rh.profiles.load_portfolio_profile("extended_hours_market_value")
        if not market_value:
            market_value = rh.profiles.load_portfolio_profile("market_value")

        if not float(market_value):
            return {}

        market_value = float(market_value)
        positions = rh.account.get_open_stock_positions()
        current_investments = {}

        for position in positions:
            symbol = rh.stocks.get_instrument_by_url(position["instrument"], info="symbol")
            available_shares = float(position["shares_available_for_exercise"])
            collateral_shares = float(position["shares_held_for_options_collateral"])
            current_price = float(rh.stocks.get_latest_price(symbol)[0])
            equity = (current_price * available_shares) + (current_price * collateral_shares)
            percentage = (equity / market_value) * 100
            current_investments[symbol] = {"equity": equity, "percentage": round(percentage, 2), "available_shares": available_shares, "collateral_shares": collateral_shares}

        return current_investments


    def get_new_investments(self):
        if platform.system().lower() == "windows":
            xlsx_file = pathlib.Path.home().joinpath("Documents", "Robinhood.xlsx")
        else:
            xlsx_file = pathlib.Path.home().joinpath("Robinhood.xlsx")

        if platform.system().lower() == "windows":
            csv_file = pathlib.Path.home().joinpath("Documents", "Robinhood.csv")
        else:
            csv_file = pathlib.Path.home().joinpath("Robinhood.csv")

        if xlsx_file.is_file():
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet = wb_obj.active
            max_row = len(sheet["A"]) - 1
            investments = {row[0].value: round(float(row[1].value), 2) for row in sheet.iter_rows(min_row=2, max_row=max_row)}
            return investments
        elif csv_file.is_file():
            _list = []
            with open(csv_file, 'r') as file:
                data = csv.reader(file)
                for row in data:
                    if row:
                        _list.append(row)
            investments = {row[0]: float(row[1]) for row in _list[1:]}
            return investments
        else:
            message = "Excel/CSV file not found. Run generate_excel_file()."
            self._log(logging.ERROR, message)
            raise Exception(message)


    def cancel_open_orders(self, side="all", sim=False):
        orders = rh.orders.get_all_open_stock_orders()

        for order in orders:
            order_id = order["id"]
            order_side = order["side"] 
            symbol = rh.stocks.get_instrument_by_url(order["instrument"], info="symbol")

            if (side == "all") or (side == order_side):
                if not sim:
                    rh.orders.cancel_stock_order(order_id)
                self._log(logging.INFO, f"Cancelling {order_side} order of {symbol}.", sim)


    def sell_all_stocks(self, sim=False):
        current_investments = self.get_current_investments()

        for instrument, data in current_investments.items():
            available_shares = current_investments[instrument]["available_shares"]
            collateral_shares = current_investments[instrument]["collateral_shares"]
            if available_shares:
                if not sim:
                    order = rh.orders.order_sell_fractional_by_quantity(instrument, available_shares, "gfd", True)
                self._log(logging.INFO, f"Selling {available_shares} share(s) of {instrument}.", sim)
            
            if collateral_shares:
                self._log(logging.INFO, f"{collateral_shares} share(s) of {instrument} held as collateral", sim)
            
            time.sleep(5)


    def rebalance_old(self, sim=False):
        self._check_list()

        buying_power = float(rh.profiles.load_account_profile(info="buying_power"))
        portfolio_value = self._get_portfolio_value()
        new_investments = self.get_new_investments()
        current_investments = self.get_current_investments()
        
        total = 0
        counter = 0 

        new_investments_set = set(new_investments.keys())
        current_investments_set = set(current_investments.keys())
        sell_list = list(current_investments_set - new_investments_set)

        for instrument in current_investments.values():
            if instrument["collateral_shares"]:
                message = "Shares held as collateral. Cannot rebalance."
                self._log(logging.ERROR, message)
                raise Exception(message)
        
        for instrument in sell_list:
            amount_to_sell = current_investments[instrument]["equity"]
            shares = current_investments[instrument]["available_shares"]
            
            if not sim:
                order = rh.orders.order_sell_fractional_by_quantity(instrument, shares, "gfd", True)
            
            self._log(logging.INFO, f"Selling ${amount_to_sell:.2f} of {instrument}.")
            time.sleep(5)
            counter += 1

            if (counter % 10) == 0:
                time.sleep(30)

        for instrument, percent in new_investments.items():
            margin_ratio = float(rh.stocks.find_instrument_data(instrument)[0]["margin_initial_ratio"])
            amount = ((buying_power + portfolio_value) * percent) / margin_ratio

            if instrument in current_investments:
                current_amount = current_investments[instrument]["equity"]
                if amount > current_amount:
                    amount_to_buy = amount - current_amount
                    
                    if not sim:
                        order = rh.orders.order_buy_fractional_by_price(instrument, amount_to_buy, "gfd", True)

                    self._log(logging.INFO, f"Buying ${amount_to_buy:.2f} from {instrument} to adjust to ${amount:.2f}")
                elif amount < current_amount:
                    amount_to_sell = current_amount - amount
                    
                    if not sim:
                        order = rh.orders.order_sell_fractional_by_price(instrument, amount_to_sell, "gfd", True)

                    self._log(logging.INFO, f"Selling ${amount_to_sell:.2f} from {instrument} to adjust to ${amount:.2f}")
                else:
                    self._log(logging.INFO, f"${amount:.2f} already invested in {instrument}")
            else:
                if not sim:
                    order = rh.orders.order_buy_fractional_by_price(instrument, amount, "gfd", True)

                self._log(logging.INFO, f"Investing ${amount:.2f} in {instrument}")
            
            time.sleep(5)
            counter += 1

            if (counter % 10) == 0:
                time.sleep(30)

            total += amount

        self._log(logging.INFO, f"Total Invested ${total:.2f}")
 

    def rebalance(self, sim=False):
        self._check_list()

        if self._check_if_collateral():
            message = "Shares held as collateral. Cannot rebalance."
            self._log(logging.ERROR, message, sim)
            raise Exception(message)

        account_type = rh.profiles.load_account_profile(info="type").lower()

        if account_type == "margin":
            self.rebalance_margin(sim)
        else:
            self.rebalance_cash(sim)


    def rebanalce_margin(self, sim=False):
        pass


    def rebanalce_cash(self, sim=False):
        new_investments = self.get_new_investments()
        
        for instrument, percent in new_investments.items():
            amount_to_invest = 4


    def test(self):
        print("Test Function Executed")